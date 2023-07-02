from adressclear import *
from openpyxl import load_workbook

if __name__ == '__main__':
    wb1 = load_workbook("518.xlsx")
    rosreestr_data = wb1["Пом2"]
    wb2 = load_workbook("home_del.xlsx")
    rosreestr_data = wb1["Пом2"]
    zhkh_data = wb2["снос по сосотоянию на 2019"]

    adr_zhkh_list = [str(zhkh_data[i][3].value) + "," +str(zhkh_data[i][4].value) for i in range(3, zhkh_data.max_row+1)]
    adr_zhkh_list_clear = []
    for adr_zhkh in adr_zhkh_list:
        adr_zhkh_clear = adressclear.get_clear_adress(adr_zhkh)
        del adr_zhkh_clear['pom']
        adr_zhkh_list_clear.append(adr_zhkh_clear)

    for i in range(2, 10):
        adress_rosreestr = rosreestr_data[i][14].value
        adress_rosreestr_clear = adressclear.get_clear_adress(adress_rosreestr)
        del adress_rosreestr_clear['pom']
        print(adress_rosreestr_clear)
        for adr_zhkh in adr_zhkh_list_clear:
            if adress_rosreestr_clear == adr_zhkh:
                print("Совпадение!!!!!!!!!!!!!", adress_rosreestr_clear)
                print("Номер строки:", i)
                rosreestr_data[i][30].value = "Снесен"
    # print(adr_ross)

    # a = adressclear.get_clear_adress("л. 40 лет ВЛКСМ,7а")
    # print(a)
    wb1.save('518.xlsx')
